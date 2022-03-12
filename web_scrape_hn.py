import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def scraping_hacker_news(*urls):
	'''
		Scrapes the "Hacker News" urls passed as arguments
		Returns two lists containing scraped html data

		Items of 'links' list contain news title and its link

		Items of 'subtexts' list contain number of votes for that news
	'''
	links, subtexts = [], []
	for x in urls:
		res = requests.get(x)
		soup = BeautifulSoup(res.text, 'html.parser')
		links.extend(soup.select('.titlelink'))
		subtexts.extend(soup.select('.subtext'))
	return links, subtexts

links, subtexts = scraping_hacker_news('https://news.ycombinator.com/news', 'https://news.ycombinator.com/news?p=2')

def sort_stories_by_votes(hn_list):
	'''
		Returns list that is sorted by number of votes
	'''
	return sorted(hn_list, key = lambda k:k[2], reverse = True)

def create_custom_hn(links, subtexts):
	'''
		Creates a list 'hn' containing lists of news title, its link and its votes

		Returns the list 'hn'
	'''
	hn = []
	for idx, item in enumerate(links):
		title = links[idx].getText()
		href = links[idx].get('href', None)
		link = subtexts[idx].select('.score')
		if link:
			points = int(link[0].getText().replace(' points', ''))
			if points > 99:
				hn.append([title, href, points])
	return sort_stories_by_votes(hn)

news = create_custom_hn(links, subtexts)

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "Top News"

# Creating Header in spreadsheet
c1 = sheet.cell(row = 1, column = 1)
c1.value = "Title"
c2 = sheet.cell(row = 1, column = 2)
c2.value = "Link"
c3 = sheet.cell(row = 1, column = 3)
c3.value = "Votes"

sheet.column_dimensions['A'].width = 60
sheet.column_dimensions['B'].width = 90

cell_alignment = Alignment(horizontal = 'general', vertical = 'center', wrap_text = True)
side_styles = Side(border_style = 'medium', color = '00000000')
all_borders = Border(top = side_styles, bottom = side_styles, left = side_styles, right = side_styles)


header_font = Font(bold = True, name = 'Times New Roman', color = '00FFFFFF', size = 14)
header_bg = PatternFill(fill_type = 'solid', start_color = '00000000', end_color = '00000000')
for j in range(3):
	each_cell = sheet.cell(row = 1, column = j+1)
	each_cell.font = header_font
	each_cell.fill = header_bg
	each_cell.border = all_borders
	each_cell.alignment = cell_alignment

# Appending the contents inside 'news' list to the sheet
for row in news:
	sheet.append(row)

rows = sheet.max_row
columns = sheet.max_column
for i in range(2, rows+1):
	for j in range(1, columns+1):
		each_cell = sheet.cell(row = i, column = j)
		if j == 2:
			link = each_cell.value
			if link.startswith('item'):
				link = 'https://news.ycombinator.com/' + link
			each_cell.value = f"=HYPERLINK(\"{link}\", \"{link}\")"
			each_cell.font = Font(size = 12, name = 'Times New Roman', color = '000000FF', italic = True, underline = 'single')
		else:
			each_cell.font = Font(size = 12, name = 'Times New Roman')
		each_cell.alignment = cell_alignment
		each_cell.border = all_borders

wb.save('Hacker News.xlsx')